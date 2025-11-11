from flask import render_template, flash, redirect, url_for, request, current_app, session
from flask_login import current_user, login_required
from app import db
from app.core import bp
import os
import openpyxl
# ¡Importante! Necesitamos 'openpyxl.utils' para la corrección
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
from app.models import Plantilla, MapaPlantilla

# Imports de Formularios (sin cambios)
from app.core.forms import (
    PlantillaUploadForm, 
    SelectSheetForm, 
    SelectHeaderRowForm, 
    MapHeadersForm
)

# Constantes (sin cambios)
ALLOWED_EXTENSIONS = {'xlsx', 'docx'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==============================================================================
# RUTA 1: DASHBOARD (Sin cambios)
# ==============================================================================
@bp.route('/dashboard', methods=['GET', 'POST'])
@login_required
def dashboard():
    """Muestra el dashboard con el formulario de subida y la lista de plantillas."""
    
    form = PlantillaUploadForm()
    
    if form.validate_on_submit():
        archivo = form.archivo_plantilla.data
        
        if archivo and allowed_file(archivo.filename):
            filename = secure_filename(archivo.filename)
            tipo_archivo = 'Excel' if filename.endswith('.xlsx') else 'Word'
            
            upload_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
            archivo.save(upload_path)
            
            nueva_plantilla = Plantilla(
                nombre_plantilla=form.nombre_plantilla.data,
                tipo_archivo=tipo_archivo,
                filename_seguro=filename,
                autor=current_user
            )
            db.session.add(nueva_plantilla)
            db.session.commit()
            
            flash(f"¡Plantilla '{nueva_plantilla.nombre_plantilla}' subida! Ahora, configura el mapeo.", "success")
            
            if tipo_archivo == 'Excel':
                return redirect(url_for('core.map_step_1_sheet', plantilla_id=nueva_plantilla.id))
            else:
                flash("El mapeo interactivo para Word aún no está implementado.", "info")
                return redirect(url_for('core.dashboard'))
        
        else:
            flash("Error: Tipo de archivo no permitido.", "danger")

    plantillas = Plantilla.query.filter_by(autor=current_user).order_by(Plantilla.timestamp.desc()).all()
    return render_template(
        'core/dashboard.html', 
        title='Dashboard', 
        form=form, 
        plantillas=plantillas
    )

# ==============================================================================
# RUTA 2: ELIMINAR PLANTILLA (Sin cambios)
# ==============================================================================
@bp.route('/delete_plantilla/<int:plantilla_id>', methods=['POST'])
@login_required
def delete_plantilla(plantilla_id):
    """Elimina una plantilla y su archivo físico."""
    
    plantilla = Plantilla.query.filter_by(id=plantilla_id, autor=current_user).first_or_404()
    
    try:
        path_archivo = os.path.join(current_app.config['UPLOAD_FOLDER'], plantilla.filename_seguro)
        if os.path.exists(path_archivo):
            os.remove(path_archivo)
            
        db.session.delete(plantilla)
        db.session.commit()
        
        flash(f"Plantilla '{plantilla.nombre_plantilla}' eliminada correctamente.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error al eliminar la plantilla: {str(e)}", "danger")
        
    return redirect(url_for('core.dashboard'))

# ==============================================================================
# RUTA 3: VER DETALLE DE PLANTILLA (Sin cambios)
# ==============================================================================
@bp.route('/ver_plantilla/<int:plantilla_id>')
@login_required
def ver_plantilla(plantilla_id):
    """Muestra el detalle del mapeo de una plantilla (lo que se guardó)."""
    
    plantilla = Plantilla.query.filter_by(id=plantilla_id, autor=current_user).first_or_404()
    mapas = plantilla.mapas.order_by(MapaPlantilla.coordenada).all()
    
    return render_template(
        'core/ver_plantilla.html', 
        title=f"Detalle: {plantilla.nombre_plantilla}", 
        plantilla=plantilla, 
        mapas=mapas
    )

# ==============================================================================
# RUTA 4: ASISTENTE PASO 1 (Sin cambios)
# ==============================================================================
@bp.route('/map_step_1_sheet/<int:plantilla_id>', methods=['GET', 'POST'])
@login_required
def map_step_1_sheet(plantilla_id):
    """Asistente - Paso 1: El usuario selecciona la hoja de Excel."""
    
    plantilla = Plantilla.query.filter_by(id=plantilla_id, autor=current_user).first_or_404()
    
    if plantilla.tipo_archivo != 'Excel':
        flash("Esta función es solo para archivos Excel.", "danger")
        return redirect(url_for('core.dashboard'))
    
    form = SelectSheetForm()
    
    try:
        path_archivo = os.path.join(current_app.config['UPLOAD_FOLDER'], plantilla.filename_seguro)
        workbook = openpyxl.load_workbook(path_archivo, read_only=True)
        sheet_names = workbook.sheetnames
        form.sheet_name.choices = [(name, name) for name in sheet_names]
        workbook.close()
        
    except Exception as e:
        flash(f"Error al leer el archivo Excel: {str(e)}", "danger")
        return redirect(url_for('core.dashboard'))
    
    if form.validate_on_submit():
        plantilla.sheet_name = form.sheet_name.data
        db.session.commit()
        return redirect(url_for('core.map_step_2_row', plantilla_id=plantilla.id))
        
    return render_template(
        'core/map_step_1_sheet.html', 
        title="Asistente (Paso 1 de 3)", 
        form=form, 
        plantilla=plantilla
    )

# ==============================================================================
# RUTA 5: ASISTENTE PASO 2 (¡CORREGIDA!)
# ==============================================================================
@bp.route('/map_step_2_row/<int:plantilla_id>', methods=['GET', 'POST'])
@login_required
def map_step_2_row(plantilla_id):
    """
    Asistente - Paso 2: El usuario selecciona la fila de encabezados
    viendo una vista previa del Excel.
    """
    
    plantilla = Plantilla.query.filter_by(id=plantilla_id, autor=current_user).first_or_404()
    
    if not plantilla.sheet_name:
        flash("Error: Primero debes seleccionar una hoja.", "danger")
        return redirect(url_for('core.map_step_1_sheet', plantilla_id=plantilla.id))
        
    form = SelectHeaderRowForm()
    
    preview_data = [] 
    row_choices = []  
    MAX_PREVIEW_ROWS = 30 
    MAX_PREVIEW_COLS = 20 

    # --- ¡INICIO DE LA CORRECCIÓN! ---
    column_headers = [] # Guardará ['A', 'B', 'C', ...]
    # --- FIN DE LA CORRECCIÓN ---

    try:
        path_archivo = os.path.join(current_app.config['UPLOAD_FOLDER'], plantilla.filename_seguro)
        workbook = openpyxl.load_workbook(path_archivo, read_only=True, data_only=True)
        sheet = workbook[plantilla.sheet_name]

        # --- ¡INICIO DE LA CORRECCIÓN! ---
        # Generar los encabezados de columna (A, B, C...)
        # Usamos 'get_column_letter' de openpyxl.utils
        for i in range(1, MAX_PREVIEW_COLS + 1):
            column_headers.append(get_column_letter(i))
        # --- FIN DE LA CORRECCIÓN! ---

        row_index = 1
        for row in sheet.iter_rows(min_row=1, max_row=MAX_PREVIEW_ROWS, max_col=MAX_PREVIEW_COLS):
            
            row_choices.append((row_index, f'Fila {row_index}'))
            cells_data = [cell.value for cell in row]
            preview_data.append(cells_data)
            row_index += 1

        workbook.close()
    
    except Exception as e:
        flash(f"Error al leer la hoja de Excel para la vista previa: {str(e)}", "danger")
        return redirect(url_for('core.map_step_1_sheet', plantilla_id=plantilla.id))

    form.header_row.choices = row_choices

    if form.validate_on_submit():
        plantilla.header_row = form.header_row.data
        db.session.commit()
        return redirect(url_for('core.map_step_3_columns', plantilla_id=plantilla.id))

    return render_template(
        'core/map_step_2_row.html', 
        title="Asistente (Paso 2 de 3)", 
        form=form, 
        plantilla=plantilla,
        preview_data=preview_data,
        column_headers=column_headers # <-- Pasamos la nueva variable al template
    )

# ==============================================================================
# RUTA 6: ASISTENTE PASO 3 (Sin cambios)
# ==============================================================================
@bp.route('/map_step_3_columns/<int:plantilla_id>', methods=['GET', 'POST'])
@login_required
def map_step_3_columns(plantilla_id):
    """Asistente - Paso 3: El usuario selecciona las columnas a mapear."""
    
    plantilla = Plantilla.query.filter_by(id=plantilla_id, autor=current_user).first_or_404()
    
    if not plantilla.sheet_name or not plantilla.header_row:
        flash("Error: Faltan pasos previos (Hoja o Fila).", "danger")
        return redirect(url_for('core.map_step_1_sheet', plantilla_id=plantilla.id))
        
    form = MapHeadersForm()
    
    headers_encontrados = {} 
    try:
        path_archivo = os.path.join(current_app.config['UPLOAD_FOLDER'], plantilla.filename_seguro)
        workbook = openpyxl.load_workbook(path_archivo, read_only=True)
        sheet = workbook[plantilla.sheet_name]
        
        for cell in sheet[plantilla.header_row]:
            if cell.value:
                headers_encontrados[cell.column_letter] = str(cell.value)
                
        workbook.close() 
                
    except Exception as e:
        flash(f"Error al leer la fila de encabezados: {str(e)}", "danger")
        return redirect(url_for('core.map_step_2_row', plantilla_id=plantilla.id))

    if not headers_encontrados:
        flash(f"No se encontraron encabezados en la fila {plantilla.header_row}. Por favor, verifica el número.", "danger")
        return redirect(url_for('core.map_step_2_row', plantilla_id=plantilla.id))

    form.headers.choices = [
        (col, f'{texto} (Col {col})') for col, texto in headers_encontrados.items()
    ]
    
    if form.validate_on_submit():
        MapaPlantilla.query.filter_by(id_plantilla=plantilla.id).delete()
        
        columnas_seleccionadas = form.headers.data
        
        for col_letra in columnas_seleccionadas:
            etiqueta_texto = headers_encontrados[col_letra]
            
            nuevo_mapa = MapaPlantilla(
                etiqueta=etiqueta_texto,
                coordenada=col_letra,
                tipo_mapa='fila_tabla',
                plantilla_padre=plantilla
            )
            db.session.add(nuevo_mapa)
            
        db.session.commit()
        
        flash("¡Mapeo completado y guardado exitosamente!", "success")
        return redirect(url_for('core.ver_plantilla', plantilla_id=plantilla.id))

    return render_template(
        'core/map_step_3_columns.html', 
        title="Asistente (Paso 3 de 3)", 
        form=form, 
        plantilla=plantilla
    )